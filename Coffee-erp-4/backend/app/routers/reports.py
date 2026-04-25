from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import get_db
from app import models
from app.auth import get_current_user

router = APIRouter()

def style_header(cell, bg="1F4E79", fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_subheader(cell):
    cell.font = Font(bold=True, name="Arial", size=9)
    cell.fill = PatternFill("solid", start_color="D6E4F0")
    cell.alignment = Alignment(horizontal="left")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

@router.get("/export/costing-report")
def export_costing_report(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    wb = openpyxl.Workbook()

    # ── Summary_Compare sheet ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary_Compare"
    ws.append(["Coffee ERP – Scenario Comparison Dashboard"])
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.append([])

    headers = ["Metric", "Fresh Cherry", "Dry Cherry", "Parchment", "Green Beans", "Best", "Worst", "Units"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=col))

    results = db.query(models.CostingResult).all()
    by_source = {r.source_type: r for r in results}
    fc = by_source.get("fresh_cherry")
    dc = by_source.get("dry_cherry")
    pa = by_source.get("parchment")
    gb = by_source.get("green_bean")

    def v(r, field):
        if r is None: return 0
        val = getattr(r, field, 0)
        return float(val) if val is not None else 0

    rows_data = [
        ("Roasted bean output (kg)", "roasted_output_kg", "kg"),
        ("Cups produced", "cups_produced", "cups"),
        ("Total variable cost", "total_variable_cost", "Rs"),
        ("Cost per roasted kg", "cost_per_roasted_kg", "Rs/kg"),
        ("Coffee cost per cup", "coffee_cost_per_cup", "Rs/cup"),
        ("Cup sales revenue", "cup_sales_revenue", "Rs"),
        ("Contribution before fixed", "contribution_before_fixed", "Rs"),
        ("Annual fixed cost", "annual_fixed_cost", "Rs"),
        ("Net profit after fixed", "net_profit", "Rs"),
        ("Net margin", "net_margin", "%"),
        ("Break-even cups/day", "breakeven_cups_per_day", "cups/day"),
    ]
    for label, field, unit in rows_data:
        vals = [v(fc, field), v(dc, field), v(pa, field), v(gb, field)]
        non_zero = [x for x in vals if x != 0]
        best = max(non_zero) if non_zero else 0
        worst = min(non_zero) if non_zero else 0
        ws.append([label] + vals + [best, worst, unit])

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ── Annual P&L sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Annual_PnL")
    ws2.append(["Annual Café P&L by Procurement Source"])
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    ws2.append([])
    pnl_headers = ["Line Item", "Fresh Cherry", "Dry Cherry", "Parchment", "Green Beans", "Units"]
    ws2.append(pnl_headers)
    for col, h in enumerate(pnl_headers, 1):
        style_header(ws2.cell(row=3, column=col))

    pnl_rows = [
        ("Cup sales revenue", "cup_sales_revenue", "Rs"),
        ("Coffee variable cost", "total_variable_cost", "Rs"),
        ("Café variable cost", "cafe_variable_cost", "Rs"),
        ("Contribution before fixed", "contribution_before_fixed", "Rs"),
        ("Fixed cost", "annual_fixed_cost", "Rs"),
        ("Net profit", "net_profit", "Rs"),
        ("Net margin", "net_margin", "%"),
    ]
    for label, field, unit in pnl_rows:
        ws2.append([label, v(fc, field), v(dc, field), v(pa, field), v(gb, field), unit])

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ── Procurement sheet ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Procurement")
    prec_headers = ["ID", "Source Type", "Supplier", "Qty (kg)", "Price/kg", "Total Cost", "Transport", "Date", "Invoice Ref"]
    ws3.append(prec_headers)
    for col, h in enumerate(prec_headers, 1):
        style_header(ws3.cell(row=1, column=col))
    procs = db.query(models.Procurement).order_by(models.Procurement.purchase_date.desc()).all()
    for p in procs:
        ws3.append([
            p.id, p.source_type, p.supplier_name,
            float(p.quantity_kg), float(p.price_per_kg),
            float(p.quantity_kg) * float(p.price_per_kg),
            float(p.transport_cost or 0),
            str(p.purchase_date), p.invoice_ref
        ])
    for col in range(1, 10):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # ── Expenses sheet ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Expenses_Ledger")
    exp_headers = ["ID", "Date", "Main Title", "Sub Title", "GL Group", "Amount (Rs)", "Reference"]
    ws4.append(exp_headers)
    for col, h in enumerate(exp_headers, 1):
        style_header(ws4.cell(row=1, column=col))
    expenses = db.query(models.Expense).order_by(models.Expense.expense_date.desc()).all()
    for e in expenses:
        ws4.append([e.id, str(e.expense_date), e.main_title, e.sub_title, e.gl_group, float(e.amount), e.reference_no])
    for col in range(1, 8):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=coffee_erp_report.xlsx"}
    )
